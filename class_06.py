# _*_ config:utf-8 _*_
# @Time     :2020/9/2 21:56
# @Author   :71_muyao
# @Email    :2646511824
# @File     :class_06
# @Software :PyCharm
'''
接口自动化的步骤：
1、提前准备好要用的接口测试用例----done
2、Python代码读取接口测试用例----done----read_data
3、读取到的用例发给requests库发送请求----done----api_request()
4、执行结果与预期结果比对====判断用例是否通过
5、比对结果回写到Excel表格里面----done---通过openpyxl实现----rewrite_result()
定义函数：
1、实现主要功能；2、对凡是变化的值定义成参数；3、返回值-----凡是别人需要从你这里得到的值；
代码自动读取测试数据+自动化回写数据====测试用例，一般都是Excel表格
第三方库：openpyxl---读取、回写
步骤：
1、安装：pip install openpyxl
2、导入
Excel常用操作：三大对象
1、工作簿---文件本身
2、表单----sheet表
3、单元格---cell
'''
#1.读取接口测试数据--Excel表格
import openpyxl
def read_data(fileName,sheetName):
    qcd_case=openpyxl.load_workbook(fileName)   #读取工作簿
    sheet=qcd_case[sheetName]   #获取到表单
    case_list=[]   #定义一个空列表
    max_row=sheet.max_row  #获取行的最大数
    for i in range(2,max_row+1):
        case=dict(
        case_id=sheet.cell(row=i,column=1).value,
        url=sheet.cell(row=i,column=5).value,  #通过行、列找到对应单元格
        data=sheet.cell(row=i,column=6).value,
        expected=sheet.cell(row=i,column=7).value
        )   #封装成大字典,一个字典是一条用例
    #print(cell.value)   #通过单元格的.value获取里面的内容
        case_list.append(case)  #大字典转换成大列表
    return case_list  #返回值
# print(cases)

#2.发送接口请求
import requests    #导入第三方库---只作用于一定范围
def api_request(api_url,api_data):
    api_headers={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  #请求头,可以导入，也可以写死，一般没有太大变化
    r=requests.post(url=api_url, json=api_data, headers=api_headers)  # 返回值，响应消息
    return r.json()   #得到响应结果

#4.结果回写
def rewrite_result(fileName,sheetName,row,column,final_result):
    qcd_case=openpyxl.load_workbook(fileName)   #读取工作簿
    sheet=qcd_case[sheetName]   #获取到表单
    sheet.cell(row=row,column=column).value=final_result    #数据重写
    qcd_case.save(fileName)   #保存回写的结果

#3.执行结果和预期结果比对
def func_data(fileName,sheetName):
    cases=read_data(fileName,sheetName)
    for j in cases:
        case_id = j["case_id"]
        url = j.get("url")
        data = j.get("data")
        data=eval(data)   #字符串---转换为---字典，eval()函数
        expected = j.get("expected")
        expected=eval(expected)
        real_result=api_request(api_url=url,api_data=data)    #执行结果
        real_msg=real_result["msg"]
        expected_msg=expected["msg"]
        print("执行结果是：{}".format(real_msg))
        print("期望结果是：{}".format(expected_msg))
        if real_msg==expected_msg:
            print("第{}条用例执行通过".format(case_id))
            final_result="Passed"
        else:
            print("第{}条用例不通过".format(case_id))
            final_result = "False"
        print("**"*20)
        rewrite_result(fileName,sheetName,case_id+1,8,final_result)
f=func_data("test_case_api.xlsx","login")
